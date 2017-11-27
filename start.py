from openpyxl import *
from collections import Counter

wb = load_workbook('SampleData.xlsx')

first_sheet = wb.get_sheet_names()[0]

ws = wb.get_sheet_by_name(first_sheet)

# all diferent values takes as keys in dict and assigns them value (how many times word appears
cnt = Counter()

for k in range(ws.min_row, ws.max_row):
    word = ws.cell(row=k+1, column=4).value
    cnt[word] += 1

new_file = Workbook()

new_ws = new_file.active

z = 0

for key, count in cnt.items():
    z = z+1
    new_ws.cell(row=z+1, column=1).value = key
    new_ws.cell(row=z + 1, column=2).value = count

new_file.save('counted_items.xlsx')


# sums total per product and adds third row to existing file

    # prepares worksheet for saving data
edit_file = load_workbook('counted_items.xlsx')

first_s = edit_file.get_sheet_names()[0]

work_sheet = edit_file.get_sheet_by_name(first_s)

# saves total per product in dict

sums = {}

for a in range(ws.min_row, ws.max_row):
    product = ws.cell(row=a+1, column=4).value
    total = ws.cell(row=a+1, column=7).value

    if product in sums:
        sums[product] = sums[product] + total

    else:
        sums[product] = total

b = 0

# saves from dict to file
for product, total in sums.items():
    b = b + 1
    if work_sheet.cell(row=b + 1, column=1).value == product:
        work_sheet.cell(row=b + 1, column=3).value = total

edit_file.save('counted_items.xlsx')
